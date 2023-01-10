import socket
import sys, os
import threading
import time
import customtkinter as ctk
import openpyxl
import resultWin

from openpyxl.utils import get_column_letter

import pickle

IP = socket.gethostbyname(socket.gethostname())
print(f"Server host : {IP}")
PORT = 5555
ADDR = (IP, PORT)

SIZE = 1024 * 2
FORMAT = 'utf-8'

FETCH = "!FETCH"
DISCONNECT_MSG = '!DISCONNECT'
RESULTS = '!RESULTS'

res = set()
cons=set()
thread_lock = threading.Lock()
control = True

l = [1, 2, 3, 4, 5, 6, 7]
question = []
A = []
B = []
C = []
D = []
ref = [A, B, C, D]
correct = []

sendList = []

resultDict={}

def loadData(path, conn, sendList):
    A.clear()
    B.clear()
    C.clear()
    D.clear()
    correct.clear()
    print(f"IN LOAD [server]")
    excelBook = openpyxl.load_workbook(path)

    ws = excelBook.active
    r = ws.rows
    checkHead = ['questions', 'question', 'a', 'b', 'c', 'd', 'answers', 'answer', '']
    num_rows = len(ws['A'])
    flag = 0
    for i in ws['A']:  # loads question
        if (str(i.value).strip().lower() in checkHead[0:2]):
            flag = 1
            continue

        elif (str(i.value).strip().lower() != "none" and flag == 1):
            flag = 1
            question.append(str(i.value).strip())

    for row in range(num_rows):  # loads options
        idx = 0
        flag = 0
        for col in range(2, 6):

            let = get_column_letter(col)

            if (str(ws[let + str(row + 1)].value).strip().lower() in checkHead[2:6]):
                flag = 1

                continue

            elif (str(ws[let + str(row + 1)].value).strip().lower() != "none"):
                flag = 1

                ref[idx].append(str(ws[let + str(row + 1)].value).strip())
                idx += 1
    # print(ref)


    flag = 0
    for i in ws['F']:  # loads correct answers
        if (str(i.value).strip().lower() in checkHead[6:8]):
            flag = 1
            continue

        elif (str(i.value).strip().lower() != "none" and flag == 1):
            flag = 1
            correct.append(str(i.value).strip())

    sendList.clear()
    sendList = [question, A, B, C, D, correct]

    with open(os.getcwd() + "\data.txt", 'w+') as file:
        file.write("")
        file.write(str(sendList))



def handle_client(conn, addr):
    print(f'[NEW CONNECTION {addr} connected]')
    connected = True
    while connected:

        msg = str(conn.recv(SIZE).decode(FORMAT))

        global action,command
        try:
            print("\nmsg from interface",msg)
            command, action = msg.split("$")

            print(f"\n[COMMAND] = {command}")


            if (command == '!LOAD'):
                loadData(action, conn, sendList)





            if (command == '!FETCH'):
                print('[IN FETCH SERVER]')
                with open(os.getcwd() + "\data.txt", 'r') as file:
                    fileData = file.read()
                    sendData = str(fileData).encode(FORMAT)
                    conn.send(sendData)





            # if(command=="!DC"):
            #     print('IN DC')
            #     with thread_lock:
            #         for i in cons:
            #             i.close()





            if (command=="!RESULTS"):
                print('[RESULT SERVER]')

                resultDict[action.split("#")[0]]=action.split("#")[1]
                with thread_lock:
                    res.add(str(resultDict))
                print(res)#here command has the clinet name and obj the result n/m
                #
                # with thread_lock:
                #     for i in cons:
                #         i.close()
                conn.close()
                break


            if (command == '!STOP'):
                print(f'[SHUT DOWN SERVER] {action}')
                nms=[]#names
                mks=[]#marks

                for i in res:
                    d=eval(i)
                    n,m=list(d.keys())[0],list(d.values())[0]
                    print(f"server name : {n} \nServer marks {m}")
                    nms.append(n)
                    mks.append(m)


                resultWin.saveResult(nms,mks,action)



                conn.close()
                os._exit(0)
        except Exception as E :

            print("Error occured IN SERVER",E,f" msg = {msg}")
            conn.close()
            break




def startServer():
    server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server.bind(ADDR)

    print(f'[Listening  server] on IP {IP} : {PORT}')
    print('[Starting server]')
    server.listen()

    while control:

        conn, addr = server.accept()
        thread = threading.Thread(target=handle_client, args=(conn, addr))

        A.clear()
        B.clear()
        C.clear()
        D.clear()
        question.clear()
        correct.clear()
        resultDict.clear()

        with thread_lock:
            cons.add(conn)
        thread.start()
        # thread.join()
        print(f'[ACTIVE CONNECTIONS] {threading.activeCount()}')

# startServer()

