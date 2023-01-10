from datetime import date
from tkinter import filedialog
import customtkinter as ctk
from openpyxl import Workbook,load_workbook
import  time
bg = "#1a1a1c"
accent = "#07d969"
text = "#fff6df"
secondary = "#c3083f"
filename_arr=['']

def save(names,marks,path,quizName):
    print("names = ",names)
    print("score = ",marks)
    root.destroy()

    name=path+"/"+quizName+"__"+str(date.today())+str(time.strftime("%H:%M:%S",time.localtime() )).replace(':','_')+'.xlsx'
    print(f"saving path {name}")
#---------------saving---------
    wb=Workbook()
    ws=wb.active
    ws.title='Results'
    wb.save(name)
# ---------------saving---------

# ---------------opening---------
    wb=load_workbook(name)
    ws=wb.active
    ws.append(["Name","Score"])
    for i,j in zip(names,marks):
        ws.append([i,j])

    wb.save(name)




    print('in save')

def open_file():
    global filename
    filename = ""
    filename = filename + str(filedialog.askdirectory())
    filename_arr[0] = filename  # stores directory address of the question file selected
    label.configure(text=filename_arr[0])
def saveResult(names, marks,quizName):
    global root
    root = ctk.CTk()
    root.geometry("500x250")
    root.title("Save results")
    global label
    label = ctk.CTkLabel(root, text="Select folder to store result file", font=("Gill Sans MT", 20))
    label.pack(pady=30)

    select=ctk.CTkButton(master=root, text="Open folder", font=("Gill Sans MT", 20), corner_radius=5,
                               command=open_file,
                               fg_color=secondary, width=100)
    select.pack(pady=20)

    save_button = ctk.CTkButton(master=root, text="Save", font=("Gill Sans MT", 20), corner_radius=5,
                           command=lambda: save(names,marks,filename_arr[0],quizName),
                           fg_color="#2c8c00", width=200)
    save_button.pack(pady=20)


    root.mainloop()




#
# saveResult(list("abcd"),[1,2,3,4],"HELLO")
print("window close")